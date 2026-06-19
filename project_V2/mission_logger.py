"""
================================================================================
  mission_logger.py  —  Mission Data Logger
================================================================================
  Version 3.0

  فولدرين بس:

  1) CAMERA DETECTIONS  — logs/camera_detections/
       صورة لكل detection جديد من الكاميرا
       - لو نفس الـ object لسه موجود → صورة واحدة بس (مش spam)
       - لو اختفى وجه تاني → صورة جديدة
       - auto-delete لما يتخطى MAX_CAMERA_IMAGES صورة

  2) OBSTACLE EVENTS  — logs/obstacle_events/
       بيتفعل فقط لما الـ Ultrasonic الأمامي يشوف عائق
       - صورة للعائق
       - Excel sheet: رقم + اسم + وقت + GPS + بيئة + صورة
       - ده اللي هيترفع على Cloud

  ملاحظة: لما Arduino مش موصل → رسالة واحدة في اللوج وكمّل طبيعي

  Author  : Robot Team
  Version : 3.0
================================================================================
"""

import cv2
import os
import threading
import time
import logging

from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from vision_processor import VisionProcessor, DetectionResult
from arduino_bridge import ArduinoBridge, Telemetry

log = logging.getLogger("MissionLogger")

# ──────────────────────────────────────────────
#  CONSTANTS
# ──────────────────────────────────────────────

CAMERA_DIR        = "logs/camera_detections"
OBSTACLE_DIR      = "logs/obstacle_events"
OBSTACLE_EXCEL    = os.path.join(OBSTACLE_DIR, "obstacles_log.xlsx")

# عدد صور الكاميرا الأقصى قبل الـ auto-delete
MAX_CAMERA_IMAGES = 100

# cooldown بين صورة وصورة لنفس الـ label في الكاميرا
CAMERA_COOLDOWN_S = 3.0

# cooldown بين event وعائق لنفس الـ label في Ultrasonic
OBSTACLE_COOLDOWN_S = 10.0

# Excel columns
EXCEL_HEADERS = [
    "Obstacle #",
    "Obstacle Name",
    "Timestamp",
    "DateTime",
    "GPS Latitude",
    "GPS Longitude",
    "Environment",
    "Image Path",
]

# ──────────────────────────────────────────────
#  GPS PLACEHOLDER
#  ← هنا هتربط GPS لما يجي من الأردوينو
#  في الـ Telemetry هتضيف: tel.gps_lat, tel.gps_lon
# ──────────────────────────────────────────────

GPS_NOT_CONNECTED = "N/A"


# ──────────────────────────────────────────────
#  DATA MODELS
# ──────────────────────────────────────────────

@dataclass
class ObstacleEvent:
    obstacle_num: int   = 0
    label:        str   = ""
    timestamp:    float = field(default_factory=time.time)
    gps_lat:      str   = GPS_NOT_CONNECTED
    gps_lon:      str   = GPS_NOT_CONNECTED
    environment:  str   = "Unknown"
    image_path:   str   = ""

    def to_excel_row(self) -> list:
        return [
            self.obstacle_num,
            self.label,
            f"{self.timestamp:.3f}",
            datetime.fromtimestamp(self.timestamp).strftime("%Y-%m-%d %H:%M:%S"),
            self.gps_lat,
            self.gps_lon,
            self.environment,
            self.image_path,
        ]


# ──────────────────────────────────────────────
#  EXCEL MANAGER
# ──────────────────────────────────────────────

class ExcelManager:
    """بيتولى إنشاء وتحديث الـ Excel file للعوائق."""

    def __init__(self, path: str):
        self.path = path
        self._lock = threading.Lock()
        self._ensure_file()

    def _ensure_file(self):
        if os.path.exists(self.path):
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Obstacle Log"

        # Header style
        header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill   = PatternFill("solid", start_color="1F4E79")
        header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border   = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        col_widths = [12, 20, 15, 22, 15, 15, 18, 40]

        for col_idx, (header, width) in enumerate(zip(EXCEL_HEADERS, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        wb.save(self.path)

    def append_row(self, event: ObstacleEvent):
        with self._lock:
            try:
                wb = load_workbook(self.path)
                ws = wb.active

                row_idx = ws.max_row + 1
                row_data = event.to_excel_row()

                data_font   = Font(name="Arial", size=10)
                data_align  = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"),
                )
                # تلوين الصفوف بالتناوب
                row_fill = PatternFill("solid", start_color="D6E4F0") \
                    if row_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font      = data_font
                    cell.alignment = data_align
                    cell.border    = thin_border
                    cell.fill      = row_fill
                    # الصورة تبقى left-aligned
                    if col_idx == len(EXCEL_HEADERS):
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                wb.save(self.path)

            except Exception as e:
                log.error(f"Excel write error: {e}")

    def reset(self):
        with self._lock:
            if os.path.exists(self.path):
                os.remove(self.path)
        self._ensure_file()


# ──────────────────────────────────────────────
#  MAIN CLASS
# ──────────────────────────────────────────────

class MissionLogger:
    """
    يراقب:
      1. الكاميرا: لو YOLO اكتشف object جديد → صورة في camera_detections
         - لو نفس الـ object لسه موجود → مش بياخد صورة تانية
         - بعد MAX_CAMERA_IMAGES صورة → بيمسح الأقدم أوتوماتيك

      2. الـ Ultrasonic (placeholder): لما يشوف عائق →
         صورة + Excel row في obstacle_events
         - بياخد GPS من الأردوينو (لو موصل)
         - بياخد البيئة الحالية من الـ vision
    """

    def __init__(
        self,
        vision: VisionProcessor,
        bridge: ArduinoBridge,
    ):
        self.vision = vision
        self.bridge = bridge

        self._running  = False
        self._thread: Optional[threading.Thread] = None
        self._lock     = threading.Lock()

        # ── Camera tracking ──────────────────────
        # label → آخر وقت حفظنا فيه
        self._camera_last_save: dict  = {}
        # label → هل لسه موجود في الـ frame
        self._camera_active:    dict  = {}
        self._camera_count:     int   = 0

        # ── Obstacle tracking ────────────────────
        self._obstacle_count:         int   = 0
        self._obstacle_last_save:     dict  = {}  # label → last time
        # ← PLACEHOLDER: لما البريدج يبعتلك إشارة Ultrasonic
        # غيّر _ultrasonic_triggered() بالـ logic الحقيقي
        self._ultrasonic_flag:        bool  = False

        # ── Arduino connection warning ────────────
        # بنطبع رسالة الـ "not connected" مرة واحدة بس
        self._arduino_warning_shown:  bool  = False

        # ── Excel ────────────────────────────────
        if not EXCEL_AVAILABLE:
            log.warning("openpyxl not installed — Excel logging disabled. "
                        "Run: pip install openpyxl")
        self._excel: Optional[ExcelManager] = \
            ExcelManager(OBSTACLE_EXCEL) if EXCEL_AVAILABLE else None

        self._setup_dirs()
        log.info("MissionLogger v3.0 initialized")

    # ──────────────────────────────────────────
    #  SETUP
    # ──────────────────────────────────────────

    def _setup_dirs(self):
        os.makedirs(CAMERA_DIR,   exist_ok=True)
        os.makedirs(OBSTACLE_DIR, exist_ok=True)
        self._camera_count = self._count_images(CAMERA_DIR)
        log.info(f"Found {self._camera_count} existing camera images")

    # ──────────────────────────────────────────
    #  PUBLIC API
    # ──────────────────────────────────────────

    def start(self):
        if self._running:
            return
        self._running = True
        self._thread  = threading.Thread(
            target=self._monitor_loop,
            daemon=True,
            name="MissionLogger",
        )
        self._thread.start()
        log.info("MissionLogger started")

    def stop(self):
        self._running = False
        if self._thread:
            self._thread.join(timeout=2.0)
        log.info("MissionLogger stopped — "
                 f"camera={self._camera_count}  obstacles={self._obstacle_count}")

    def stats(self) -> dict:
        return {
            "camera_images":    self._camera_count,
            "obstacles_logged": self._obstacle_count,
            "camera_dir_mb":    self._dir_size_mb(CAMERA_DIR),
            "obstacle_dir_mb":  self._dir_size_mb(OBSTACLE_DIR),
        }

    def trigger_ultrasonic_obstacle(self, label: str = "obstacle"):
        """
        ← استدعي الـ method دي من الأردوينو بريدج لما الـ Ultrasonic يشوف عائق.
        مثال في arduino_bridge.py:
            if ultrasonic_front < THRESHOLD:
                self._mission_logger.trigger_ultrasonic_obstacle(label="object")
        """
        with self._lock:
            self._ultrasonic_flag  = True
            self._ultrasonic_label = label

    # ──────────────────────────────────────────
    #  CLEAR
    # ──────────────────────────────────────────

    def clear_camera(self):
        count = self._count_images(CAMERA_DIR)
        size  = self._dir_size_mb(CAMERA_DIR)
        self._clear_dir_images(CAMERA_DIR)
        self._camera_count = 0
        self._camera_last_save.clear()
        self._camera_active.clear()
        log.info(f"Camera log cleared — was {count} images ({size:.1f} MB)")

    def clear_obstacles(self):
        count = self._count_images(OBSTACLE_DIR)
        size  = self._dir_size_mb(OBSTACLE_DIR)
        self._clear_dir_images(OBSTACLE_DIR)
        if self._excel:
            self._excel.reset()
        self._obstacle_count = 0
        self._obstacle_last_save.clear()
        log.info(f"Obstacle log cleared — was {count} images ({size:.1f} MB)")

    def clear_all(self):
        self.clear_camera()
        self.clear_obstacles()
        log.info("All logs cleared")

    # ──────────────────────────────────────────
    #  MONITOR LOOP
    # ──────────────────────────────────────────

    def _monitor_loop(self):
        interval = 0.1  # 10 Hz

        while self._running:
            t0 = time.time()
            try:
                self._tick()
            except Exception as e:
                log.error(f"MissionLogger tick error: {e}", exc_info=True)

            elapsed = time.time() - t0
            sleep_t = interval - elapsed
            if sleep_t > 0:
                time.sleep(sleep_t)

    def _tick(self):
        # ── 1. Camera detection ───────────────
        result = self.vision.get_latest()
        self._handle_camera(result)

        # ── 2. Ultrasonic obstacle ────────────
        with self._lock:
            triggered      = self._ultrasonic_flag
            obstacle_label = getattr(self, "_ultrasonic_label", "obstacle")
            self._ultrasonic_flag = False  # reset بعد ما قريناه

        if triggered:
            self._handle_obstacle_event(obstacle_label)

    # ──────────────────────────────────────────
    #  CAMERA LOGGING
    # ──────────────────────────────────────────

    def _handle_camera(self, result: DetectionResult):
        """
        لو في detection:
          - لو نفس الـ label لسه شايفه → ما يحفظش
          - لو اختفى وجه تاني → يحفظ
          - لو جديد تماماً → يحفظ
        """
        if not result.obstacle_found() or not result.is_valid(max_age=0.5):
            # مفيش detection — نعمل reset للـ active labels
            with self._lock:
                self._camera_active.clear()
            return

        label = result.label
        now   = time.time()

        with self._lock:
            was_active = self._camera_active.get(label, False)
            self._camera_active[label] = True

            if was_active:
                # نفس الـ object لسه موجود — مش محتاج صورة
                return

            last_save = self._camera_last_save.get(label, 0.0)
            if now - last_save < CAMERA_COOLDOWN_S:
                return

            self._camera_last_save[label] = now

        frame = self.vision.get_raw_frame()

        threading.Thread(
            target=self._save_camera_image,
            args=(frame, label, now),
            daemon=True,
            name="CameraSave",
        ).start()

    def _save_camera_image(self, frame, label: str, timestamp: float):
        try:
            ts   = datetime.fromtimestamp(timestamp).strftime("%Y%m%d_%H%M%S")
            name = f"CAM_{ts}_{label}.jpg"
            path = os.path.join(CAMERA_DIR, name)

            if frame is not None:
                cv2.imwrite(path, frame)
                self._camera_count += 1
                log.debug(f"Camera image saved: {name} (total={self._camera_count})")
                self._enforce_camera_limit()
            else:
                log.debug(f"Camera detection ({label}) — no frame available")

        except Exception as e:
            log.error(f"Camera save error: {e}")

    def _enforce_camera_limit(self):
        """لو بقى أكتر من MAX_CAMERA_IMAGES → امسح الأقدم."""
        try:
            images = sorted([
                os.path.join(CAMERA_DIR, f)
                for f in os.listdir(CAMERA_DIR)
                if f.lower().endswith((".jpg", ".jpeg", ".png"))
            ])

            while len(images) > MAX_CAMERA_IMAGES:
                oldest = images.pop(0)
                os.remove(oldest)
                log.debug(f"Auto-deleted old camera image: {os.path.basename(oldest)}")

        except Exception as e:
            log.error(f"Camera limit enforcement error: {e}")

    # ──────────────────────────────────────────
    #  OBSTACLE EVENT LOGGING
    # ──────────────────────────────────────────

    def _handle_obstacle_event(self, label: str):
        """
        بيتفعل لما الـ Ultrasonic يشوف عائق.
        بياخد: صورة + GPS + بيئة → يحفظ في obstacle_events
        """
        now = time.time()

        # cooldown للـ label ده
        last_save = self._obstacle_last_save.get(label, 0.0)
        if now - last_save < OBSTACLE_COOLDOWN_S:
            return

        self._obstacle_last_save[label] = now
        self._obstacle_count += 1
        num = self._obstacle_count

        # GPS
        gps_lat, gps_lon = self._get_gps()

        # البيئة الحالية
        env = self.vision.get_environment()
        env_label = env.label if env and env.label not in ("", "Unknown") else "Unknown"

        # اسم العائق من الـ YOLO لو موجود
        yolo_result = self.vision.get_latest()
        if yolo_result.obstacle_found() and yolo_result.is_valid(max_age=1.0):
            obstacle_name = yolo_result.label
        else:
            obstacle_name = label  # fallback للـ label اللي جاء من الـ Ultrasonic

        log.info(
            f"OBSTACLE EVENT #{num} | name={obstacle_name} | "
            f"env={env_label} | GPS=({gps_lat}, {gps_lon})"
        )

        frame = self.vision.get_raw_frame()

        event = ObstacleEvent(
            obstacle_num = num,
            label        = obstacle_name,
            timestamp    = now,
            gps_lat      = gps_lat,
            gps_lon      = gps_lon,
            environment  = env_label,
        )

        threading.Thread(
            target=self._save_obstacle_event,
            args=(frame, event),
            daemon=True,
            name="ObstacleSave",
        ).start()

    def _save_obstacle_event(self, frame, event: ObstacleEvent):
        try:
            ts   = datetime.fromtimestamp(event.timestamp).strftime("%Y%m%d_%H%M%S")
            name = f"OBS_{event.obstacle_num:04d}_{ts}_{event.label}.jpg"
            path = os.path.join(OBSTACLE_DIR, name)

            if frame is not None:
                cv2.imwrite(path, frame)
                event.image_path = path
                log.debug(f"Obstacle image saved: {name}")
            else:
                event.image_path = "NO_FRAME"
                log.warning(f"Obstacle #{event.obstacle_num} — no frame available")

            # Excel
            if self._excel:
                self._excel.append_row(event)
            else:
                log.warning("Excel not available — obstacle logged without Excel row")

        except Exception as e:
            log.error(f"Obstacle event save error: {e}")

    # ──────────────────────────────────────────
    #  GPS HELPER
    # ──────────────────────────────────────────

    def _get_gps(self) -> Tuple[str, str]:
        """
        ← PLACEHOLDER: هنا هتجيب الـ GPS من الـ Telemetry
        لما تضيف GPS للـ Telemetry في arduino_bridge.py:
            tel = self.bridge.get_telemetry()
            return str(tel.gps_lat), str(tel.gps_lon)
        """
        try:
            tel = self.bridge.get_telemetry()

            # ← لما تضيف GPS للـ Telemetry uncommint الـ lines دول:
            # if hasattr(tel, 'gps_lat') and tel.gps_lat is not None:
            #     return str(tel.gps_lat), str(tel.gps_lon)

            # Arduino مش موصل أو مفيش GPS — رسالة واحدة بس
            if not self._arduino_warning_shown:
                log.warning("Arduino not connected — GPS will be logged as N/A")
                self._arduino_warning_shown = True

            return GPS_NOT_CONNECTED, GPS_NOT_CONNECTED

        except Exception:
            if not self._arduino_warning_shown:
                log.warning("Arduino not connected — GPS will be logged as N/A")
                self._arduino_warning_shown = True
            return GPS_NOT_CONNECTED, GPS_NOT_CONNECTED

    # ──────────────────────────────────────────
    #  HELPERS
    # ──────────────────────────────────────────

    @staticmethod
    def _count_images(directory: str) -> int:
        if not os.path.exists(directory):
            return 0
        return sum(
            1 for f in os.listdir(directory)
            if f.lower().endswith((".jpg", ".jpeg", ".png"))
        )

    @staticmethod
    def _clear_dir_images(directory: str):
        if not os.path.exists(directory):
            return
        for f in os.listdir(directory):
            if f.lower().endswith((".jpg", ".jpeg", ".png", ".xlsx")):
                try:
                    os.remove(os.path.join(directory, f))
                except Exception as e:
                    log.warning(f"Could not delete {f}: {e}")

    @staticmethod
    def _dir_size_mb(directory: str) -> float:
        if not os.path.exists(directory):
            return 0.0
        total = sum(
            os.path.getsize(os.path.join(directory, f))
            for f in os.listdir(directory)
            if os.path.isfile(os.path.join(directory, f))
        )
        return total / (1024 * 1024)


# ──────────────────────────────────────────────
#  QUICK TEST
# ──────────────────────────────────────────────

if __name__ == "__main__":
    import logging
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(name)s] %(levelname)s — %(message)s",
    )

    from arduino_bridge import create_bridge

    print("=" * 60)
    print("  MissionLogger v3.0 — Simulation Test")
    print("=" * 60)

    bridge = create_bridge(simulate=True)
    vision = VisionProcessor(simulate=True, use_yolo=False, use_env=False)

    bridge.start()
    vision.start()

    logger = MissionLogger(vision, bridge)
    logger.start()

    # ── Test 1: Camera detection ──────────────
    print("\n[TEST 1] Camera detection — same object repeated")
    vision._latest_result = DetectionResult(
        position="FORWARD", label="bottle",
        confidence=0.85, threat_level="LOW",
        approx_dist="MEDIUM", persistent=True,
        timestamp=time.time(),
    )
    time.sleep(1.5)
    # نفس الـ object لسه موجود — مش المفروض ياخد صورة تانية
    time.sleep(1.5)
    vision._latest_result = DetectionResult()  # اختفى
    time.sleep(0.5)
    # ظهر تاني — المفروض ياخد صورة جديدة
    vision._latest_result = DetectionResult(
        position="FORWARD", label="bottle",
        confidence=0.80, threat_level="LOW",
        approx_dist="MEDIUM", persistent=True,
        timestamp=time.time(),
    )
    time.sleep(1.0)
    vision._latest_result = DetectionResult()
    print(f"  Camera images: {logger._camera_count} (expected: 2)")

    # ── Test 2: Ultrasonic trigger ────────────
    print("\n[TEST 2] Ultrasonic obstacle trigger")
    logger.trigger_ultrasonic_obstacle(label="wall")
    time.sleep(0.5)
    print(f"  Obstacles logged: {logger._obstacle_count} (expected: 1)")

    # ── Test 3: Ultrasonic cooldown ───────────
    print("\n[TEST 3] Ultrasonic cooldown — same obstacle twice fast")
    logger.trigger_ultrasonic_obstacle(label="wall")
    time.sleep(0.3)
    logger.trigger_ultrasonic_obstacle(label="wall")
    time.sleep(0.5)
    print(f"  Obstacles logged: {logger._obstacle_count} (expected: still 1)")

    # ── Stats ─────────────────────────────────
    print(f"\n  Stats: {logger.stats()}")

    # ── Clear ────────────────────────────────
    print("\n[TEST 4] Clear all")
    logger.clear_all()
    print(f"  After clear — Stats: {logger.stats()}")

    logger.stop()
    vision.stop()
    bridge.stop()

    print("\n" + "=" * 60)
    print("  Done ✅")
    print("=" * 60)