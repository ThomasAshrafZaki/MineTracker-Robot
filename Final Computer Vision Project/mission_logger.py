"""
================================================================================
  mission_logger.py  —  Mission Data Logger
================================================================================
  Version 3.0

  فولدرين بس:

  1) CAMERA DETECTIONS  — logs/camera_detections/
       صورة لكل detection جديد من الكاميرا
       - لو نفس الـ object لسه موجود → صورة واحدة بس (مش spam)
       - لو اختفى وجه تاني → صورة جديدة
       - auto-delete لما يتخطى MAX_CAMERA_IMAGES صورة

  2) OBSTACLE EVENTS  — logs/obstacle_events/
       بيتفعل فقط لما الـ Ultrasonic الأمامي يشوف عائق
       - صورة للعائق
       - Excel sheet: رقم + اسم + وقت + GPS + بيئة + صورة
       - ده اللي هيترفع على Cloud

  ملاحظة: لما Arduino مش موصل → رسالة واحدة في اللوج وكمّل طبيعي

  ── v3.1 — Camera logging fixes ──────────────────────────────────
    ✓ FIX 1: الصورة المحفوظة بقت annotated frame (بالمربع والـ HUD)
      بدل raw frame — بنستخدم get_frame() بدل get_raw_frame().
    ✓ FIX 2: بطلنا نحفظ أي صورة أو Excel row لو label فاضي — لازم
      يكون في اسم detection حقيقي صريح قبل أي تسجيل.
    ✓ FIX 3: ضفنا miss-streak debounce (CAMERA_MISS_STREAK_NEEDED)
      عشان ما نعتبرش الجسم اختفى إلا بعد 1.5 ثانية غياب فعلي —
      بيمنع الـ spam بسبب تذبذب YOLO المؤقت.

  Author  : Robot Team
  Version : 3.1
================================================================================
"""

import cv2
import os
import threading
import time
import logging

from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from vision_processor import VisionProcessor, DetectionResult
from arduino_bridge import ArduinoBridge, Telemetry

log = logging.getLogger("MissionLogger")

# ──────────────────────────────────────────────
#  CONSTANTS
# ──────────────────────────────────────────────

CAMERA_DIR        = "logs/camera_detections"
CAMERA_EXCEL      = os.path.join(CAMERA_DIR,   "camera_detections_log.xlsx")
OBSTACLE_DIR      = "logs/obstacle_events"
OBSTACLE_EXCEL    = os.path.join(OBSTACLE_DIR, "obstacles_log.xlsx")

# عدد صور الكاميرا الأقصى قبل الـ auto-delete
MAX_CAMERA_IMAGES = 100

# cooldown بين صورة وصورة لنفس الـ label في الكاميرا
CAMERA_COOLDOWN_S = 3.0

# ── FIX v3.1: miss-streak debounce ──────────────────────────────
# كام تيك متتالي بدون detection قبل ما نعتبر الجسم "اختفى فعليًا"
# كل تيك = 100ms (10Hz) → 15 تيك = 1.5 ثانية
# بيمنع تكرار الصور بسبب تذبذب YOLO المؤقت (مثلاً بطء الـ inference)
CAMERA_MISS_STREAK_NEEDED = 15

# ── v3.2 — Stability + Sharpness gating ──────────────────────────
# (بيتحكموا بس في *الحفظ*، مش في الـ detection نفسه — الـ detection
# بتاع vision_processor.py فضل زي ما هو من غير أي تعديل)

# كام تيك متتالي لازم يفضل نفس الـ label موجود (وهو detection حقيقي
# valid) قبل ما نعتبره "مستقر" ونبدأ نفكر نصوره أصلاً.
# 4 تيك ≈ 400ms — بيمنع إننا نصور لقطة لحظية (موبايل اتشاف غلط كإنسان
# لمدة فريم واحد مثلاً) قبل ما تتأكد كفاية.
CAMERA_STABLE_HITS_NEEDED = 4

# الثقة المطلوبة وقت *الحفظ* — أعلى من اللي بيتعرض على الشاشة عمداً.
# ممكن تتعرض على الشاشة من ثقة أقل، بس مش هتتسجل في اللوج/الصور إلا
# لو الثقة وصلت للحد ده وقت محاولة الحفظ.
CAMERA_SAVE_MIN_CONF = 0.60

# أقل قيمة Laplacian variance عشان نعتبر الفريم "واضح" مش مهزوز/مشوش.
# لو الفريم مش واضح، منحفظوش فورًا — بنستنى محاولات تانية.
CAMERA_SHARPNESS_MIN_VAR = 60.0

# أقصى عدد تيكات (محاولات) نستنى فيها فريم واضح وثقته كافية، قبل ما
# نستسلم ونحفظ أحسن فريم لقيناه لحد دلوقتي (عشان منضيعش الحدث خالص).
# 10 تيك ≈ 1 ثانية.
CAMERA_CAPTURE_MAX_ATTEMPTS = 10

# نفس فكرة الـ sharpness بس لصور العوائق — عدد محاولات بسيطة جوه
# نفس الـ save-thread (مش بيوقف أي حاجة تانية، الـ thread أصلاً مستقل).
OBSTACLE_SHARPNESS_MIN_VAR  = 50.0
OBSTACLE_CAPTURE_ATTEMPTS   = 4
OBSTACLE_CAPTURE_RETRY_GAP  = 0.05  # ثانية بين كل محاولة

# cooldown بين event وعائق لنفس الـ label في Ultrasonic
OBSTACLE_COOLDOWN_S = 10.0

# ── Mine detection logging (v3.2) ────────────────────────────────
# مفيش فولدر ولا Excel منفصل للغم — بيتسجل في camera_detections نفسها
# بالظبط زي أي detection عادي (نفس الفولدر، نفس الإكسل، نفس الترقيم).
# الـ cooldown هنا بس عشان لو نفس اللغم لسه قدام الكاميرا والموديل
# بيأكد عليه كل تيك، متتسجلش صور مكررة كل 100ms.
MINE_LOG_COOLDOWN_S = 5.0

# Obstacle Excel columns
EXCEL_HEADERS = [
    "Obstacle #",
    "Obstacle Name",
    "Timestamp",
    "DateTime",
    "GPS Latitude",
    "GPS Longitude",
    "Image Path",
]

# Camera Excel columns
CAMERA_EXCEL_HEADERS = [
    "Detection #",
    "Detection Name",
    "Timestamp",
    "DateTime",
    "Image Name",
]

# ──────────────────────────────────────────────
#  GPS PLACEHOLDER
#  ← هنا هتربط GPS لما يجي من الأردوينو
#  في الـ Telemetry هتضيف: tel.gps_lat, tel.gps_lon
# ──────────────────────────────────────────────

GPS_NOT_CONNECTED = "N/A"


# ──────────────────────────────────────────────
#  DATA MODELS
# ──────────────────────────────────────────────

@dataclass
class CameraDetectionEvent:
    detection_num: int   = 0
    label:         str   = ""
    timestamp:     float = field(default_factory=time.time)
    image_name:    str   = ""

    def to_excel_row(self) -> list:
        return [
            self.detection_num,
            self.label,
            f"{self.timestamp:.3f}",
            datetime.fromtimestamp(self.timestamp).strftime("%Y-%m-%d %H:%M:%S"),
            self.image_name,
        ]


@dataclass
class ObstacleEvent:
    obstacle_num: int   = 0
    label:        str   = ""
    timestamp:    float = field(default_factory=time.time)
    gps_lat:      str   = GPS_NOT_CONNECTED
    gps_lon:      str   = GPS_NOT_CONNECTED
    image_path:   str   = ""

    def to_excel_row(self) -> list:
        return [
            self.obstacle_num,
            self.label,
            f"{self.timestamp:.3f}",
            datetime.fromtimestamp(self.timestamp).strftime("%Y-%m-%d %H:%M:%S"),
            self.gps_lat,
            self.gps_lon,
            self.image_path,
        ]


# ──────────────────────────────────────────────
#  EXCEL MANAGER
# ──────────────────────────────────────────────

class ExcelManager:
    """بيتولى إنشاء وتحديث الـ Excel file للعوائق."""

    def __init__(self, path: str):
        self.path = path
        self._lock = threading.Lock()
        self._ensure_file()

    def _ensure_file(self):
        if os.path.exists(self.path):
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Obstacle Log"

        # Header style
        header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill   = PatternFill("solid", start_color="1F4E79")
        header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border   = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        col_widths = [12, 20, 15, 22, 15, 15, 40]

        for col_idx, (header, width) in enumerate(zip(EXCEL_HEADERS, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        wb.save(self.path)

    def append_row(self, event: ObstacleEvent):
        with self._lock:
            try:
                wb = load_workbook(self.path)
                ws = wb.active

                row_idx = ws.max_row + 1
                row_data = event.to_excel_row()

                data_font   = Font(name="Arial", size=10)
                data_align  = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"),
                )
                # تلوين الصفوف بالتناوب
                row_fill = PatternFill("solid", start_color="D6E4F0") \
                    if row_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font      = data_font
                    cell.alignment = data_align
                    cell.border    = thin_border
                    cell.fill      = row_fill
                    # الصورة تبقى left-aligned
                    if col_idx == len(EXCEL_HEADERS):
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                wb.save(self.path)

            except Exception as e:
                log.error(f"Excel write error: {e}")

    def reset(self):
        with self._lock:
            if os.path.exists(self.path):
                os.remove(self.path)
        self._ensure_file()


class CameraExcelManager:
    """Excel log للـ camera detections — بيتراكم ومش بيتمسح منه صفوف."""

    def __init__(self, path: str):
        self.path  = path
        self._lock = threading.Lock()
        self._detection_count = 0
        self._ensure_file()

    def _ensure_file(self):
        if os.path.exists(self.path):
            # احسب عدد الصفوف الموجودة عشان الترقيم يكمل صح
            try:
                wb = load_workbook(self.path, read_only=True)
                ws = wb.active
                self._detection_count = max(0, ws.max_row - 1)
                wb.close()
            except Exception:
                self._detection_count = 0
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Camera Detections"

        header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill("solid", start_color="1A5276")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        col_widths = [14, 22, 15, 22, 35]

        for col_idx, (header, width) in enumerate(zip(CAMERA_EXCEL_HEADERS, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        wb.save(self.path)

    def append_row(self, event: CameraDetectionEvent):
        with self._lock:
            try:
                self._detection_count += 1
                event.detection_num = self._detection_count

                wb = load_workbook(self.path)
                ws = wb.active

                row_idx     = ws.max_row + 1
                data_font   = Font(name="Arial", size=10)
                data_align  = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"),
                )
                row_fill = PatternFill("solid", start_color="D5E8D4") \
                    if row_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")

                for col_idx, value in enumerate(event.to_excel_row(), start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font      = data_font
                    cell.alignment = data_align
                    cell.border    = thin_border
                    cell.fill      = row_fill

                wb.save(self.path)

            except Exception as e:
                log.error(f"Camera Excel write error: {e}")

    def reset(self):
        with self._lock:
            if os.path.exists(self.path):
                os.remove(self.path)
            self._detection_count = 0
        self._ensure_file()


# ──────────────────────────────────────────────
#  MAIN CLASS
# ──────────────────────────────────────────────

class MissionLogger:
    """
    يراقب:
      1. الكاميرا: لو YOLO اكتشف object جديد → صورة في camera_detections
         - لو نفس الـ object لسه موجود → مش بياخد صورة تانية
         - بعد MAX_CAMERA_IMAGES صورة → بيمسح الأقدم أوتوماتيك

      2. الـ Ultrasonic (placeholder): لما يشوف عائق →
         صورة + Excel row في obstacle_events
         - بياخد GPS من الأردوينو (لو موصل)
         - بياخد البيئة الحالية من الـ vision
    """

    def __init__(
        self,
        vision: VisionProcessor,
        bridge: ArduinoBridge,
    ):
        self.vision = vision
        self.bridge = bridge

        self._running  = False
        self._thread: Optional[threading.Thread] = None
        self._lock     = threading.Lock()

        # ── Camera tracking ──────────────────────
        # label → آخر وقت حفظنا فيه
        self._camera_last_save: dict  = {}
        # ── FIX v3.1: بدل dict بيتمسح فورًا، بقى عندنا "آخر label نشط"
        # + عداد تيكات الغياب المتتالية — بيمنع spam تصوير بسبب تذبذب YOLO
        self._camera_active_label: Optional[str] = None
        self._camera_miss_streak:  int = 0
        self._camera_count:     int   = 0
        self._camera_detection_num: int = 0
        # ── v3.2: stability+sharpness capture state machine ──────
        # None = مفيش حاجة بنستقر عليها دلوقتي.
        # dict = candidate بنستنى يثبت/نلاقي فريم واضح ليه.
        self._camera_pending: Optional[dict] = None

        # ── Obstacle tracking ────────────────────
        self._obstacle_count:         int   = 0
        self._obstacle_last_save:     dict  = {}  # label → last time
        # ← PLACEHOLDER: لما البريدج يبعتلك إشارة Ultrasonic
        # غيّر _ultrasonic_triggered() بالـ logic الحقيقي
        self._ultrasonic_flag:        bool  = False

        # ── Mine detection logging (بيتسجل في camera_detections نفسها) ──
        self._mine_last_save:         float = 0.0

        # ── Arduino connection warning ────────────
        # بنطبع رسالة الـ "not connected" مرة واحدة بس
        self._arduino_warning_shown:  bool  = False

        # ── Dirs first, then Excel ────────────────
        self._setup_dirs()

        # ── Excel ────────────────────────────────
        if not EXCEL_AVAILABLE:
            log.warning("openpyxl not installed — Excel logging disabled. "
                        "Run: pip install openpyxl")
        self._excel: Optional[ExcelManager] = \
            ExcelManager(OBSTACLE_EXCEL) if EXCEL_AVAILABLE else None
        self._camera_excel: Optional[CameraExcelManager] = \
            CameraExcelManager(CAMERA_EXCEL) if EXCEL_AVAILABLE else None

        log.info("MissionLogger v3.1 initialized")

    # ──────────────────────────────────────────
    #  SETUP
    # ──────────────────────────────────────────

    def _setup_dirs(self):
        os.makedirs(CAMERA_DIR,   exist_ok=True)
        os.makedirs(OBSTACLE_DIR, exist_ok=True)
        self._camera_count = self._count_images(CAMERA_DIR)
        log.info(f"Found {self._camera_count} existing camera images")

    # ──────────────────────────────────────────
    #  PUBLIC API
    # ──────────────────────────────────────────

    def start(self):
        if self._running:
            return
        self._running = True
        self._thread  = threading.Thread(
            target=self._monitor_loop,
            daemon=True,
            name="MissionLogger",
        )
        self._thread.start()
        log.info("MissionLogger started")

    def stop(self):
        self._running = False
        if self._thread:
            self._thread.join(timeout=2.0)
        log.info("MissionLogger stopped — "
                 f"camera={self._camera_count}  obstacles={self._obstacle_count}")

    def stats(self) -> dict:
        return {
            "camera_images":    self._camera_count,
            "obstacles_logged": self._obstacle_count,
            "camera_dir_mb":    self._dir_size_mb(CAMERA_DIR),
            "obstacle_dir_mb":  self._dir_size_mb(OBSTACLE_DIR),
        }

    def trigger_ultrasonic_obstacle(self, label: str = "obstacle"):
        """
        ← استدعي الـ method دي من الأردوينو بريدج لما الـ Ultrasonic يشوف عائق.
        مثال في arduino_bridge.py:
            if ultrasonic_front < THRESHOLD:
                self._mission_logger.trigger_ultrasonic_obstacle(label="object")
        """
        with self._lock:
            self._ultrasonic_flag  = True
            self._ultrasonic_label = label

    def log_mine_detection(self, label: str, confidence: float):
        """
        ← استدعيها من main.py لما mine_detector.py (موديل best.pt الثانوي)
        يأكد كشف لغم.

        مهم: مفيش فولدر ولا Excel منفصل للغم — بتتسجل بالظبط زي أي
        camera detection عادي:
          - نفس الفولدر  logs/camera_detections/
          - نفس الإكسل   camera_detections_log.xlsx
          - Detection Name = اسم الكلاس بتاع اللغم (label)
          - الصورة = annotated frame اللي فيها صندوق اللغم (الماجنتا)
            مرسوم عليها بالفعل من vision_processor._draw_annotations

        فيها cooldown بسيط (MINE_LOG_COOLDOWN_S) بس عشان منكررش نفس
        الصورة كل تيك لو اللغم لسه قدام الكاميرا — مش منطق منفصل،
        مجرد حماية من السبام.
        """
        now = time.time()
        if now - self._mine_last_save < MINE_LOG_COOLDOWN_S:
            return
        self._mine_last_save = now

        detection_label = (label or "mine").strip() or "mine"
        frame = self.vision.get_frame()   # annotated — فيها صندوق اللغم بالفعل

        log.warning(
            f"MINE DETECTION → camera_detections | label={detection_label} "
            f"conf={confidence:.0%}"
        )

        threading.Thread(
            target=self._save_camera_image,
            args=(frame, detection_label, now),
            daemon=True,
            name="MineCameraSave",
        ).start()

    # ──────────────────────────────────────────
    #  CLEAR
    # ──────────────────────────────────────────

    def clear_camera(self):
        count = self._count_images(CAMERA_DIR)
        size  = self._dir_size_mb(CAMERA_DIR)
        self._clear_dir_images(CAMERA_DIR)
        if self._camera_excel:
            self._camera_excel.reset()
        self._camera_count = 0
        self._camera_detection_num = 0
        self._camera_last_save.clear()
        self._camera_active_label = None
        self._camera_miss_streak  = 0
        self._camera_pending      = None
        log.info(f"Camera log cleared — was {count} images ({size:.1f} MB)")

    def clear_obstacles(self):
        count = self._count_images(OBSTACLE_DIR)
        size  = self._dir_size_mb(OBSTACLE_DIR)
        self._clear_dir_images(OBSTACLE_DIR)
        if self._excel:
            self._excel.reset()
        self._obstacle_count = 0
        self._obstacle_last_save.clear()
        log.info(f"Obstacle log cleared — was {count} images ({size:.1f} MB)")

    def clear_all(self):
        self.clear_camera()
        self.clear_obstacles()
        log.info("All logs cleared")

    # ──────────────────────────────────────────
    #  MONITOR LOOP
    # ──────────────────────────────────────────

    def _monitor_loop(self):
        interval = 0.1  # 10 Hz

        while self._running:
            t0 = time.time()
            try:
                self._tick()
            except Exception as e:
                log.error(f"MissionLogger tick error: {e}", exc_info=True)

            elapsed = time.time() - t0
            sleep_t = interval - elapsed
            if sleep_t > 0:
                time.sleep(sleep_t)

    def _tick(self):
        # ── 1. Camera detection ───────────────
        result = self.vision.get_latest()
        self._handle_camera(result)

        # ── 2. Ultrasonic obstacle ────────────
        with self._lock:
            triggered      = self._ultrasonic_flag
            obstacle_label = getattr(self, "_ultrasonic_label", "obstacle")
            self._ultrasonic_flag = False  # reset بعد ما قريناه

        if triggered:
            self._handle_obstacle_event(obstacle_label)

    # ──────────────────────────────────────────
    #  CAMERA LOGGING
    # ──────────────────────────────────────────

    def _handle_camera(self, result: DetectionResult):
        """
        ── FIX v3.1 (3 مشاكل حُلّت) ──────────────────────────────────

        FIX 1 — صورة فاضية في الـ label:
          قبل كده كان بيكفي obstacle_found()==True لو في label فاضي
          بسبب persistence carry-over. دلوقتي label لازم يكون غير
          فاضي صراحة قبل أي تصوير أو تسجيل في الإكسل.

        FIX 2 — Spam تصوير بسبب تذبذب YOLO:
          بدل ما نمسح الـ active tracking فورًا لما obstacle_found()
          يرجع False (ممكن يكون سببه بطء inference مؤقت)، دلوقتي
          بنستنى CAMERA_MISS_STREAK_NEEDED تيك متتالي (1.5 ثانية)
          قبل ما نعتبر الجسم اختفى فعليًا.

        FIX 3 — الصورة بتتحفظ raw بدون detection overlay:
          get_raw_frame() اتستبدلت بـ get_frame() عشان الصورة
          المحفوظة تكون نفس اللي بتتعرض على الشاشة (بالمربع والـ HUD).
        """
        # ── FIX 1: لازم label حقيقي غير فاضي + detection valid + box فعلي ──
        # has_box() بيتأكد إن في مربع رسم فعلي على الـ frame (v9.3 DetectionResult)
        # بدون box = persistence carry-over بدون object حقيقي = لا تصوير ولا تسجيل
        has_valid_label = bool(result.label and result.label.strip())
        is_valid_detection = (
            result.obstacle_found()
            and result.is_valid(max_age=0.5)
            and has_valid_label
            and result.has_box()   # ← FIX v9.3: لازم bounding box فعلي
        )

        now         = time.time()
        should_save = False
        label       = ""
        frame_to_save = None

        with self._lock:
            if not is_valid_detection:
                # ── FIX 2: miss-streak debounce — مش بنمسح فورًا ────
                self._camera_miss_streak += 1
                # ── v3.2: أي غياب فعلي (حتى قبل ما الـ active label يتمسح)
                # بيبطّل أي محاولة استقرار شغالة، عشان "الاستقرار" يبقى
                # معناه فعلاً وجود متواصل بدون انقطاع
                self._camera_pending = None
                if self._camera_miss_streak >= CAMERA_MISS_STREAK_NEEDED:
                    self._camera_active_label = None
                return

            # في detection حقيقي وله اسم — نصفّر عداد الغياب
            self._camera_miss_streak = 0
            label = result.label

            if label == self._camera_active_label:
                # نفس الـ object لسه موجود — مش محتاج صورة تانية
                self._camera_pending = None
                return

            # label اتغير: object جديد أو رجع بعد غياب حقيقي كافي
            last_save = self._camera_last_save.get(label, 0.0)
            if now - last_save < CAMERA_COOLDOWN_S:
                # رجع بسرعة قبل انتهاء الـ cooldown — نعامله كـ active
                # بدون تصوير عشان منكررش المحاولة كل تيك
                self._camera_active_label = label
                self._camera_pending = None
                return

            # ── v3.2 STEP A: Stability gate ─────────────────────
            # قبل أي تفكير في الحفظ، لازم نفس الـ label يفضل ظاهر
            # CAMERA_STABLE_HITS_NEEDED تيك متتالي من غير انقطاع.
            if self._camera_pending is None or self._camera_pending["label"] != label:
                self._camera_pending = {
                    "label":           label,
                    "hits":            1,
                    "attempts":        0,
                    "best_frame":      None,
                    "best_sharpness":  -1.0,
                }
                return

            pending = self._camera_pending
            pending["hits"] += 1

            if pending["hits"] < CAMERA_STABLE_HITS_NEEDED:
                # لسه مستقرش كفاية — منصورش لسه
                return

            # ── v3.2 STEP B: Save-confidence gate ───────────────
            # ثقة الحفظ أعلى من ثقة العرض على الشاشة عمداً
            if result.confidence < CAMERA_SAVE_MIN_CONF:
                pending["attempts"] += 1
                if pending["attempts"] < CAMERA_CAPTURE_MAX_ATTEMPTS:
                    return
                # خلصت المحاولات والثقة لسه واطية — لو معندناش أي فريم
                # محفوظ من قبل، نسيب الحدث ده يضيع بدل ما نسجل حاجة
                # ضعيفة الثقة؛ لو عندنا فريم سابق (نادر هنا) نستخدمه.
                if pending["best_frame"] is None:
                    self._camera_pending = None
                    return

            # ── v3.2 STEP C: Sharpness gate ──────────────────────
            frame_now = self.vision.get_frame()
            sharpness = self._compute_sharpness(frame_now) if frame_now is not None else -1.0
            pending["attempts"] += 1

            if sharpness > pending["best_sharpness"]:
                pending["best_sharpness"] = sharpness
                pending["best_frame"]     = frame_now

            sharp_enough   = sharpness >= CAMERA_SHARPNESS_MIN_VAR
            attempts_done  = pending["attempts"] >= CAMERA_CAPTURE_MAX_ATTEMPTS

            if not sharp_enough and not attempts_done:
                # نديله فرصة تيك تاني يلاقي فريم أوضح
                return

            # ── FINALIZE ──────────────────────────────────────────
            frame_to_save = pending["best_frame"]
            self._camera_pending = None

            if frame_to_save is None:
                # ولا فريم واحد اتسجل (نادر) — متسجلش حاجة فاضية
                return

            self._camera_active_label     = label
            self._camera_last_save[label] = now
            should_save = True

        if not should_save:
            return

        threading.Thread(
            target=self._save_camera_image,
            args=(frame_to_save, label, now),
            daemon=True,
            name="CameraSave",
        ).start()

    def _save_camera_image(self, frame, label: str, timestamp: float):
        try:
            ts   = datetime.fromtimestamp(timestamp).strftime("%Y%m%d_%H%M%S")
            name = f"CAM_{ts}_{label}.jpg"
            path = os.path.join(CAMERA_DIR, name)

            if frame is not None:
                cv2.imwrite(path, frame)
                self._camera_count += 1
                log.debug(f"Camera image saved: {name} (total={self._camera_count})")
                self._enforce_camera_limit()
            else:
                log.debug(f"Camera detection ({label}) — no frame available")

            # Excel — بيتراكم ومش بيتمسح
            if self._camera_excel:
                event = CameraDetectionEvent(
                    label       = label,
                    timestamp   = timestamp,
                    image_name  = name if frame is not None else "NO_FRAME",
                )
                self._camera_excel.append_row(event)

        except Exception as e:
            log.error(f"Camera save error: {e}")

    def _enforce_camera_limit(self):
        """لو بقى أكتر من MAX_CAMERA_IMAGES → امسح الأقدم."""
        try:
            images = sorted([
                os.path.join(CAMERA_DIR, f)
                for f in os.listdir(CAMERA_DIR)
                if f.lower().endswith((".jpg", ".jpeg", ".png"))
            ])

            while len(images) > MAX_CAMERA_IMAGES:
                oldest = images.pop(0)
                os.remove(oldest)
                log.debug(f"Auto-deleted old camera image: {os.path.basename(oldest)}")

        except Exception as e:
            log.error(f"Camera limit enforcement error: {e}")

    # ──────────────────────────────────────────
    #  OBSTACLE EVENT LOGGING
    # ──────────────────────────────────────────

    def _handle_obstacle_event(self, label: str):
        """
        بيتفعل لما الـ Ultrasonic يشوف عائق.
        بياخد: صورة + GPS → يحفظ في obstacle_events
        """
        now = time.time()

        # cooldown للـ label ده
        last_save = self._obstacle_last_save.get(label, 0.0)
        if now - last_save < OBSTACLE_COOLDOWN_S:
            return

        self._obstacle_last_save[label] = now
        self._obstacle_count += 1
        num = self._obstacle_count

        # GPS
        gps_lat, gps_lon = self._get_gps()

        # اسم العائق من الـ YOLO لو موجود
        yolo_result = self.vision.get_latest()
        if yolo_result.obstacle_found() and yolo_result.is_valid(max_age=1.0):
            obstacle_name = yolo_result.label
        else:
            obstacle_name = label  # fallback للـ label اللي جاء من الـ Ultrasonic

        log.info(
            f"OBSTACLE EVENT #{num} | name={obstacle_name} | "
            f"GPS=({gps_lat}, {gps_lon})"
        )

        frame = self.vision.get_frame()   # annotated (بالمربع والـ HUD) زي الكاميرا

        event = ObstacleEvent(
            obstacle_num = num,
            label        = obstacle_name,
            timestamp    = now,
            gps_lat      = gps_lat,
            gps_lon      = gps_lon,
        )

        threading.Thread(
            target=self._save_obstacle_event,
            args=(frame, event),
            daemon=True,
            name="ObstacleSave",
        ).start()

    def _save_obstacle_event(self, frame, event: ObstacleEvent):
        try:
            # ── v3.2: obstacle events مهمة (بتترفع Cloud) — هنا إحنا
            # جوه thread مستقل بالفعل (مش الـ monitor loop)، فمسموح
            # نستنى كذا محاولة قصيرة (sleep) عشان نمسك فريم أوضح لو
            # الأول كان مهزوز، من غير ما نأثر على أي حاجة تانية شغالة.
            best_frame     = frame
            best_sharpness = self._compute_sharpness(frame) if frame is not None else -1.0

            if best_sharpness < OBSTACLE_SHARPNESS_MIN_VAR:
                for _ in range(OBSTACLE_CAPTURE_ATTEMPTS - 1):
                    time.sleep(OBSTACLE_CAPTURE_RETRY_GAP)
                    candidate  = self.vision.get_frame()
                    cand_sharp = self._compute_sharpness(candidate) if candidate is not None else -1.0
                    if cand_sharp > best_sharpness:
                        best_sharpness = cand_sharp
                        best_frame     = candidate
                    if best_sharpness >= OBSTACLE_SHARPNESS_MIN_VAR:
                        break

            frame = best_frame

            ts   = datetime.fromtimestamp(event.timestamp).strftime("%Y%m%d_%H%M%S")
            name = f"OBS_{event.obstacle_num:04d}_{ts}_{event.label}.jpg"
            path = os.path.join(OBSTACLE_DIR, name)

            if frame is not None:
                cv2.imwrite(path, frame)
                event.image_path = path
                log.debug(f"Obstacle image saved: {name} (sharpness={best_sharpness:.0f})")
            else:
                event.image_path = "NO_FRAME"
                log.warning(f"Obstacle #{event.obstacle_num} — no frame available")

            # Excel
            if self._excel:
                self._excel.append_row(event)
            else:
                log.warning("Excel not available — obstacle logged without Excel row")

        except Exception as e:
            log.error(f"Obstacle event save error: {e}")

    # ──────────────────────────────────────────
    #  GPS HELPER
    # ──────────────────────────────────────────

    def _get_gps(self) -> Tuple[str, str]:
        """
        ← PLACEHOLDER: هنا هتجيب الـ GPS من الـ Telemetry
        لما تضيف GPS للـ Telemetry في arduino_bridge.py:
            tel = self.bridge.get_telemetry()
            return str(tel.gps_lat), str(tel.gps_lon)
        """
        try:
            tel = self.bridge.get_telemetry()

            # ← لما تضيف GPS للـ Telemetry uncommint الـ lines دول:
            # if hasattr(tel, 'gps_lat') and tel.gps_lat is not None:
            #     return str(tel.gps_lat), str(tel.gps_lon)

            # Arduino مش موصل أو مفيش GPS — رسالة واحدة بس
            if not self._arduino_warning_shown:
                log.warning("Arduino not connected — GPS will be logged as N/A")
                self._arduino_warning_shown = True

            return GPS_NOT_CONNECTED, GPS_NOT_CONNECTED

        except Exception:
            if not self._arduino_warning_shown:
                log.warning("Arduino not connected — GPS will be logged as N/A")
                self._arduino_warning_shown = True
            return GPS_NOT_CONNECTED, GPS_NOT_CONNECTED

    # ──────────────────────────────────────────
    #  HELPERS
    # ──────────────────────────────────────────

    @staticmethod
    def _compute_sharpness(frame) -> float:
        """
        Laplacian variance — مقياس بسيط وسريع لوضوح الصورة.
        رقم واطي = صورة مهزوزة/مشوشة/blurred. رقم عالي = حواف واضحة.
        مش مكلف حسابيًا (frame واحد بالـ grayscale)، مناسب يتنادى كل تيك.
        """
        try:
            if frame is None:
                return -1.0
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            return float(cv2.Laplacian(gray, cv2.CV_64F).var())
        except Exception:
            return -1.0

    @staticmethod
    def _count_images(directory: str) -> int:
        if not os.path.exists(directory):
            return 0
        return sum(
            1 for f in os.listdir(directory)
            if f.lower().endswith((".jpg", ".jpeg", ".png"))
        )

    @staticmethod
    def _clear_dir_images(directory: str):
        if not os.path.exists(directory):
            return
        for f in os.listdir(directory):
            if f.lower().endswith((".jpg", ".jpeg", ".png", ".xlsx")):
                try:
                    os.remove(os.path.join(directory, f))
                except Exception as e:
                    log.warning(f"Could not delete {f}: {e}")

    @staticmethod
    def _dir_size_mb(directory: str) -> float:
        if not os.path.exists(directory):
            return 0.0
        total = sum(
            os.path.getsize(os.path.join(directory, f))
            for f in os.listdir(directory)
            if os.path.isfile(os.path.join(directory, f))
        )
        return total / (1024 * 1024)


# ──────────────────────────────────────────────
#  QUICK TEST
# ──────────────────────────────────────────────

if __name__ == "__main__":
    import logging
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(name)s] %(levelname)s — %(message)s",
    )

    from arduino_bridge import create_bridge

    print("=" * 60)
    print("  MissionLogger v3.0 — Simulation Test")
    print("=" * 60)

    bridge = create_bridge(simulate=True)
    vision = VisionProcessor(simulate=True, use_yolo=False)

    bridge.start()
    vision.start()

    logger = MissionLogger(vision, bridge)
    logger.start()

    # ── Test 1: Camera detection ──────────────
    print("\n[TEST 1] Camera detection — same object repeated")
    vision._latest_result = DetectionResult(
        position="FORWARD", label="bottle",
        confidence=0.85, threat_level="LOW",
        approx_dist="MEDIUM", persistent=True,
        timestamp=time.time(),
    )
    time.sleep(1.5)
    # نفس الـ object لسه موجود — مش المفروض ياخد صورة تانية
    time.sleep(1.5)
    vision._latest_result = DetectionResult()  # اختفى
    time.sleep(0.5)
    # ظهر تاني — المفروض ياخد صورة جديدة
    vision._latest_result = DetectionResult(
        position="FORWARD", label="bottle",
        confidence=0.80, threat_level="LOW",
        approx_dist="MEDIUM", persistent=True,
        timestamp=time.time(),
    )
    time.sleep(1.0)
    vision._latest_result = DetectionResult()
    print(f"  Camera images: {logger._camera_count} (expected: 2)")

    # ── Test 2: Ultrasonic trigger ────────────
    print("\n[TEST 2] Ultrasonic obstacle trigger")
    logger.trigger_ultrasonic_obstacle(label="wall")
    time.sleep(0.5)
    print(f"  Obstacles logged: {logger._obstacle_count} (expected: 1)")

    # ── Test 3: Ultrasonic cooldown ───────────
    print("\n[TEST 3] Ultrasonic cooldown — same obstacle twice fast")
    logger.trigger_ultrasonic_obstacle(label="wall")
    time.sleep(0.3)
    logger.trigger_ultrasonic_obstacle(label="wall")
    time.sleep(0.5)
    print(f"  Obstacles logged: {logger._obstacle_count} (expected: still 1)")

    # ── Stats ─────────────────────────────────
    print(f"\n  Stats: {logger.stats()}")

    # ── Clear ────────────────────────────────
    print("\n[TEST 4] Clear all")
    logger.clear_all()
    print(f"  After clear — Stats: {logger.stats()}")

    logger.stop()
    vision.stop()
    bridge.stop()

    print("\n" + "=" * 60)
    print("  Done ✅")
    print("=" * 60)